"""
create_test_submission.py
Generates a filled-in copy of FinSAC_Template_MP.xlsx with realistic test data
so we can run the full pipeline end-to-end.
Run once:  python create_test_submission.py
"""

import shutil
import openpyxl
from datetime import date, datetime
from pathlib import Path

SRC  = '/Users/samuelkolbeck/Desktop/ai/world bank/FinSAC_Template_MP.xlsx'
DEST = 'submissions/BGR_2026-H1_MP.xlsx'

Path('submissions').mkdir(exist_ok=True)
shutil.copy(SRC, DEST)

wb = openpyxl.load_workbook(DEST)

# ── 01_Submission_Metadata  (values go in col C = index 3) ──────────────
ws = wb['01_Submission_Metadata']
ws.cell(5,  3).value = 'BGR'
ws.cell(6,  3).value = '2026-H1'
ws.cell(7,  3).value = 2025
ws.cell(8,  3).value = date(2026, 7, 1)
ws.cell(9,  3).value = 'Test Analyst'
ws.cell(10, 3).value = 'analyst@test.bg'
ws.cell(11, 3).value = 'Test submission — pipeline smoke test'

# ── 06_Data_National  ───────────────────────────────────────────────────
# Columns: round_id, country_code, year, indicator_id, value, unit,
#          currency, fx_rate_to_eur, fx_rate_date, source,
#          is_estimated, is_forecast, reported_level, notes
ws = wb['06_Data_National']

national_rows = [
    # (indicator_id, value, unit, source)
    ('GHG_EMISSIONS_EXCL_LULUCF_KT_CO2_EQ',     47_500,  'kt CO2 eq',   'UNFCCC'),
    ('GHG_EMISSIONS_BY_GAS_CARBON_DIOXIDE_CO2_', 38_200,  'kt CO2 eq',   'UNFCCC'),
    ('GHG_EMISSIONS_BY_GAS_METHANE_CH4_KT_CO2_',  4_300,  'kt CO2 eq',   'UNFCCC'),
    ('GHG_EMISSIONS_BY_GAS_NITROUS_OXIDE_N2O_K',  2_800,  'kt CO2 eq',   'UNFCCC'),
    ('GDP_CURRENT_PRICES_USD_MN',               101_800,  'USD mn',      'World Bank'),
    ('GDP_CURRENT_PRICES_EUR_MN',                93_200,  'EUR mn',      'World Bank'),
    ('GENERAL_GOVERNMENT_GROSS_DEBT_PCT_GDP',       23.5,  '% of GDP',   'IMF'),
    ('GENERAL_GOVERNMENT_FISCAL_BALANCE_PCT_GDP',   -1.8,  '% of GDP',   'IMF'),
    ('CURRENT_ACCOUNT_BALANCE_PCT_GDP',              0.3,  '% of GDP',   'IMF'),
    ('GROSS_INTERNATIONAL_RESERVES_MONTHS_IMPORT',   6.2,  'months',     'BNB'),
]

for r_idx, (ind_id, val, unit, src) in enumerate(national_rows, 2):
    ws.cell(r_idx, 1).value  = '2026-H1'
    ws.cell(r_idx, 2).value  = 'BGR'
    ws.cell(r_idx, 3).value  = 2025
    ws.cell(r_idx, 4).value  = ind_id
    ws.cell(r_idx, 5).value  = val
    ws.cell(r_idx, 6).value  = unit
    ws.cell(r_idx, 7).value  = None        # currency
    ws.cell(r_idx, 8).value  = None        # fx_rate_to_eur
    ws.cell(r_idx, 9).value  = None        # fx_rate_date
    ws.cell(r_idx, 10).value = src
    ws.cell(r_idx, 11).value = False       # is_estimated
    ws.cell(r_idx, 12).value = False       # is_forecast
    ws.cell(r_idx, 13).value = 'national total'
    ws.cell(r_idx, 14).value = None        # notes

# ── 07_Data_Regional  ───────────────────────────────────────────────────
# Columns: round_id, country_code, year, indicator_id, region_id, value,
#          unit, currency, fx_rate_to_eur, fx_rate_date, source,
#          is_estimated, is_forecast, reported_level, notes
ws = wb['07_Data_Regional']

regional_rows = [
    ('BGR-SOFIA',  'GHG_EMISSIONS_EXCL_LULUCF_KT_CO2_EQ', 12_400, 'kt CO2 eq'),
    ('BGR-PLOVDI', 'GHG_EMISSIONS_EXCL_LULUCF_KT_CO2_EQ',  7_200, 'kt CO2 eq'),
    ('BGR-VARNA',  'GHG_EMISSIONS_EXCL_LULUCF_KT_CO2_EQ',  5_100, 'kt CO2 eq'),
    ('BGR-SOFIA',  'GDP_CURRENT_PRICES_EUR_MN',             38_500, 'EUR mn'),
    ('BGR-PLOVDI', 'GDP_CURRENT_PRICES_EUR_MN',             14_200, 'EUR mn'),
]

for r_idx, (region_id, ind_id, val, unit) in enumerate(regional_rows, 2):
    ws.cell(r_idx, 1).value  = '2026-H1'
    ws.cell(r_idx, 2).value  = 'BGR'
    ws.cell(r_idx, 3).value  = 2025
    ws.cell(r_idx, 4).value  = ind_id
    ws.cell(r_idx, 5).value  = region_id
    ws.cell(r_idx, 6).value  = val
    ws.cell(r_idx, 7).value  = unit
    ws.cell(r_idx, 8).value  = None
    ws.cell(r_idx, 9).value  = None
    ws.cell(r_idx, 10).value = None
    ws.cell(r_idx, 11).value = 'NUTS/NSI'
    ws.cell(r_idx, 12).value = False
    ws.cell(r_idx, 13).value = False
    ws.cell(r_idx, 14).value = 'NUTS2'
    ws.cell(r_idx, 15).value = None

# ── 08_Data_Sectoral  ───────────────────────────────────────────────────
ws = wb['08_Data_Sectoral']

sectoral_rows = [
    ('N0001', 'GHG_EMISSIONS_EXCL_LULUCF_KT_CO2_EQ',  8_200, 'kt CO2 eq'),  # Agriculture
    ('N0005', 'GHG_EMISSIONS_EXCL_LULUCF_KT_CO2_EQ', 14_100, 'kt CO2 eq'),  # Mining
    ('N0010', 'GHG_EMISSIONS_EXCL_LULUCF_KT_CO2_EQ', 11_300, 'kt CO2 eq'),  # Manufacturing
    ('N0001', 'GDP_CURRENT_PRICES_EUR_MN',              4_200, 'EUR mn'),
    ('N0005', 'GDP_CURRENT_PRICES_EUR_MN',              3_800, 'EUR mn'),
    ('N0010', 'GDP_CURRENT_PRICES_EUR_MN',             18_600, 'EUR mn'),
]

for r_idx, (nace_id, ind_id, val, unit) in enumerate(sectoral_rows, 2):
    ws.cell(r_idx, 1).value  = '2026-H1'
    ws.cell(r_idx, 2).value  = 'BGR'
    ws.cell(r_idx, 3).value  = 2025
    ws.cell(r_idx, 4).value  = ind_id
    ws.cell(r_idx, 5).value  = nace_id
    ws.cell(r_idx, 6).value  = val
    ws.cell(r_idx, 7).value  = unit
    ws.cell(r_idx, 8).value  = None
    ws.cell(r_idx, 9).value  = None
    ws.cell(r_idx, 10).value = None
    ws.cell(r_idx, 11).value = 'Eurostat'
    ws.cell(r_idx, 12).value = False
    ws.cell(r_idx, 13).value = False
    ws.cell(r_idx, 14).value = 'NACE rev.2'
    ws.cell(r_idx, 15).value = None

wb.save(DEST)
print(f"✅  Test file saved: {DEST}")
print(f"   National rows : {len(national_rows)}")
print(f"   Regional rows : {len(regional_rows)}")
print(f"   Sectoral rows : {len(sectoral_rows)}")
