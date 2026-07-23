"""
stages/ingest.py — Parse FinSAC Template MP Excel files (9-sheet format).

Expected sheets:
  00_README              (skipped)
  01_Submission_Metadata
  02_Ref_Country
  03_Ref_Indicator
  04_Ref_Region
  05_Ref_Sector_NACE
  06_Data_National
  07_Data_Regional
  08_Data_Sectoral
"""

import openpyxl
from datetime import datetime
from typing import Dict, List, Tuple, Optional

from models import (
    CountrySubmission, CountryMetadata, DataRow,
    RegionalDataRow, SectoralDataRow,
)


class FinSACIngestor:
    """Parse a single FinSAC _MP Excel file into a CountrySubmission."""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.errors: List[Dict] = []
        self.ref_data: Dict = {}

    def ingest(self) -> Tuple[Optional[CountrySubmission], List[Dict]]:
        try:
            self._load_ref_country()
            self._load_ref_indicator()
            self._load_ref_region()
            self._load_ref_sector_nace()

            metadata = self._parse_metadata()
            if not metadata:
                return None, self.errors

            national_data  = self._parse_national_data(metadata)
            regional_data  = self._parse_regional_data(metadata)
            sectoral_data  = self._parse_sectoral_data(metadata)

            submission = CountrySubmission(
                metadata=metadata,
                national_data=national_data,
                regional_data=regional_data,
                sectoral_data=sectoral_data,
                submission_file=self.excel_path,
            )
            return submission, self.errors

        except Exception as exc:
            self.errors.append({
                'stage': 'ingest', 'severity': 'error',
                'message': f'Exception during ingest: {exc}',
            })
            return None, self.errors

    # ------------------------------------------------------------------
    # Reference loaders
    # ------------------------------------------------------------------

    def _load_ref_country(self):
        ws = self._sheet('02_Ref_Country')
        if ws is None:
            return
        countries = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code, name, region, eu = (row + (None,) * 4)[:4]
            if not code:
                continue
            countries[str(code).strip().upper()] = {
                'name': name, 'finsac_region': region, 'eu_member': eu,
            }
        self.ref_data['countries'] = countries

    def _load_ref_indicator(self):
        ws = self._sheet('03_Ref_Indicator')
        if ws is None:
            return
        indicators = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ind_id, name, unit, module, theme, grain = (row + (None,) * 6)[:6]
            if not ind_id:
                continue
            indicators[str(ind_id).strip()] = {
                'name': name, 'unit': unit, 'module': module,
                'theme': theme, 'expected_grain': grain,
            }
        self.ref_data['indicators'] = indicators

    def _load_ref_region(self):
        ws = self._sheet('04_Ref_Region')
        if ws is None:
            return
        regions = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            region_id, country_code, region_name, level, geo = (row + (None,) * 5)[:5]
            if not region_id:
                continue
            regions[str(region_id).strip()] = {
                'country_code': country_code,
                'region_name': region_name,
                'admin_level': level,
                'geo_code': geo,
            }
        self.ref_data['regions'] = regions

    def _load_ref_sector_nace(self):
        ws = self._sheet('05_Ref_Sector_NACE')
        if ws is None:
            return
        sectors = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            node_id, nace_code, level, parent, label = (row + (None,) * 5)[:5]
            if not node_id:
                continue
            sectors[str(node_id).strip()] = {
                'nace_code': nace_code, 'nace_level': level,
                'parent_node_id': parent, 'sector_label': label,
            }
        self.ref_data['sectors'] = sectors

    # ------------------------------------------------------------------
    # Metadata
    # ------------------------------------------------------------------

    def _parse_metadata(self) -> Optional[CountryMetadata]:
        ws = self._sheet('01_Submission_Metadata', required=True)
        if ws is None:
            return None
        try:
            country_code      = ws.cell(5, 2).value
            round_id          = ws.cell(6, 2).value
            data_as_of_year   = ws.cell(7, 2).value
            submission_date_v = ws.cell(8, 2).value
            analyst_name      = ws.cell(9, 2).value
            analyst_email     = ws.cell(10, 2).value
            notes             = ws.cell(11, 2).value

            if isinstance(submission_date_v, str):
                submission_date = datetime.strptime(submission_date_v, '%Y-%m-%d')
            elif isinstance(submission_date_v, datetime):
                submission_date = submission_date_v
            else:
                submission_date = datetime.now()

            return CountryMetadata(
                country_code=str(country_code).strip().upper(),
                round_id=str(round_id).strip(),
                data_as_of_year=int(data_as_of_year) if data_as_of_year else datetime.now().year,
                submission_date=submission_date,
                analyst_name=str(analyst_name) if analyst_name else "",
                analyst_email=str(analyst_email) if analyst_email else "",
                notes=str(notes) if notes else None,
            )
        except Exception as exc:
            self.errors.append({
                'stage': 'ingest', 'severity': 'error',
                'message': f'Error parsing metadata: {exc}',
            })
            return None

    # ------------------------------------------------------------------
    # Data sheets
    # ------------------------------------------------------------------

    def _parse_national_data(self, meta: CountryMetadata) -> List[DataRow]:
        ws = self._sheet('06_Data_National')
        if ws is None:
            return []
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            _, country_code, year, indicator_id, value, unit, \
                currency, fx_rate, fx_date, source, is_estimated = (row + (None,) * 11)[:11]
            if not indicator_id or year is None:
                continue
            try:
                rows.append(DataRow(
                    country_code=str(country_code or meta.country_code).strip().upper(),
                    year=int(year),
                    indicator_id=str(indicator_id).strip(),
                    value=float(value) if value is not None else 0.0,
                    unit=str(unit) if unit else "",
                    currency=str(currency).strip().upper() if currency else None,
                    fx_rate_to_eur=float(fx_rate) if fx_rate else None,
                    fx_rate_date=str(fx_date) if fx_date else None,
                    source=str(source) if source else "",
                    is_estimated=bool(is_estimated) if is_estimated else False,
                ))
            except (ValueError, AssertionError) as exc:
                self.errors.append({
                    'stage': 'ingest', 'severity': 'warning',
                    'sheet': '06_Data_National', 'row': row_idx,
                    'message': f'Skipping malformed row: {exc}',
                })
        return rows

    def _parse_regional_data(self, meta: CountryMetadata) -> List[RegionalDataRow]:
        ws = self._sheet('07_Data_Regional')
        if ws is None:
            return []
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            _, country_code, year, indicator_id, region_id, value, unit, \
                currency, fx_rate, fx_date, source, is_estimated = (row + (None,) * 12)[:12]
            if not indicator_id or year is None or not region_id:
                continue
            try:
                rows.append(RegionalDataRow(
                    country_code=str(country_code or meta.country_code).strip().upper(),
                    region_id=str(region_id).strip(),
                    year=int(year),
                    indicator_id=str(indicator_id).strip(),
                    value=float(value) if value is not None else 0.0,
                    unit=str(unit) if unit else "",
                    currency=str(currency).strip().upper() if currency else None,
                    fx_rate_to_eur=float(fx_rate) if fx_rate else None,
                    fx_rate_date=str(fx_date) if fx_date else None,
                    source=str(source) if source else "",
                    is_estimated=bool(is_estimated) if is_estimated else False,
                ))
            except (ValueError, AssertionError) as exc:
                self.errors.append({
                    'stage': 'ingest', 'severity': 'warning',
                    'sheet': '07_Data_Regional', 'row': row_idx,
                    'message': f'Skipping malformed row: {exc}',
                })
        return rows

    def _parse_sectoral_data(self, meta: CountryMetadata) -> List[SectoralDataRow]:
        ws = self._sheet('08_Data_Sectoral')
        if ws is None:
            return []
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            _, country_code, year, indicator_id, nace_node_id, value, unit, \
                currency, fx_rate, fx_date, source, is_estimated = (row + (None,) * 12)[:12]
            if not indicator_id or year is None or not nace_node_id:
                continue
            try:
                rows.append(SectoralDataRow(
                    country_code=str(country_code or meta.country_code).strip().upper(),
                    nace_node_id=str(nace_node_id).strip(),
                    year=int(year),
                    indicator_id=str(indicator_id).strip(),
                    value=float(value) if value is not None else 0.0,
                    unit=str(unit) if unit else "",
                    currency=str(currency).strip().upper() if currency else None,
                    fx_rate_to_eur=float(fx_rate) if fx_rate else None,
                    fx_rate_date=str(fx_date) if fx_date else None,
                    source=str(source) if source else "",
                    is_estimated=bool(is_estimated) if is_estimated else False,
                ))
            except (ValueError, AssertionError) as exc:
                self.errors.append({
                    'stage': 'ingest', 'severity': 'warning',
                    'sheet': '08_Data_Sectoral', 'row': row_idx,
                    'message': f'Skipping malformed row: {exc}',
                })
        return rows

    # ------------------------------------------------------------------
    # Helper
    # ------------------------------------------------------------------

    def _sheet(self, name: str, required: bool = False):
        if name not in self.wb.sheetnames:
            if required:
                self.errors.append({
                    'stage': 'ingest', 'severity': 'error',
                    'message': f'Missing required sheet: {name}',
                })
            return None
        return self.wb[name]


def ingest(excel_path: str) -> Tuple[Optional[CountrySubmission], List[Dict]]:
    """Main entry point: parse a FinSAC _MP Excel file."""
    return FinSACIngestor(excel_path).ingest()
